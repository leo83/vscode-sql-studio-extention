"""Excel export via openpyxl."""

from __future__ import annotations

from pathlib import Path
from typing import Any, Sequence

from openpyxl import Workbook
from openpyxl.utils import get_column_letter


def export_xlsx(
    path: str | Path,
    columns: Sequence[str],
    rows: Sequence[Sequence[Any]],
) -> int:
    file_path = Path(path)
    file_path.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    ws = wb.active
    ws.title = "Results"
    ws.append(list(columns))
    for row in rows:
        ws.append(["" if v is None else v for v in row])
    ws.freeze_panes = "A2"
    for idx, col in enumerate(columns, start=1):
        letter = get_column_letter(idx)
        lengths = [len(str(col))]
        for row in rows[:100]:
            if idx - 1 < len(row) and row[idx - 1] is not None:
                lengths.append(len(str(row[idx - 1])))
        max_len = max(lengths) if lengths else 10
        ws.column_dimensions[letter].width = min(max_len + 2, 50)
    wb.save(file_path)
    return len(rows)
