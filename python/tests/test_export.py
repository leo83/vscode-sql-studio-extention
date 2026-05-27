"""Tests for CSV and Excel export."""

from __future__ import annotations

from pathlib import Path

from openpyxl import load_workbook

from sql_studio.export.csv_export import export_csv
from sql_studio.export.excel_export import export_xlsx


def test_export_csv_writes_header_and_rows(tmp_path: Path) -> None:
    path = tmp_path / "nested" / "data.csv"
    count = export_csv(path, ["id", "name"], [[1, "alice"], [2, None]], bom=True)
    assert count == 2
    text = path.read_text(encoding="utf-8-sig")
    lines = text.strip().splitlines()
    assert lines[0] == "id,name"
    assert lines[1] == "1,alice"
    assert lines[2] == "2,"


def test_export_csv_without_bom(tmp_path: Path) -> None:
    path = tmp_path / "plain.csv"
    export_csv(path, ["x"], [[1]], bom=False)
    raw = path.read_bytes()
    assert not raw.startswith(b"\xef\xbb\xbf")


def test_export_xlsx_creates_workbook(tmp_path: Path) -> None:
    path = tmp_path / "results.xlsx"
    count = export_xlsx(path, ["n"], [[1], [2], [3]])
    assert count == 3
    wb = load_workbook(path)
    ws = wb.active
    assert ws.title == "Results"
    assert ws["A1"].value == "n"
    assert ws["A2"].value == 1
    assert ws["A4"].value == 3
    assert ws.freeze_panes == "A2"
